#!/usr/bin/env python3
"""
数据汇总脚本 - 合并多次爬取的数据到一个Excel文件
"""
import os
import sys
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config_changsha as config

# 样式定义
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')


def get_all_crawl_dirs(base_dir="data/crawl_results"):
    """获取所有爬取结果目录"""
    if not os.path.exists(base_dir):
        print(f"目录不存在: {base_dir}")
        return []
    
    dirs = []
    for name in os.listdir(base_dir):
        dir_path = os.path.join(base_dir, name)
        if os.path.isdir(dir_path):
            excel_file = os.path.join(dir_path, "数据统计.xlsx")
            if os.path.exists(excel_file):
                dirs.append((dir_path, name))
    
    # 按时间戳排序（目录名格式：关键字_YYYYMMDD_HHMMSS）
    dirs.sort(key=lambda x: x[1])
    return dirs


def load_crawl_data(excel_file):
    """加载单个爬取结果的数据"""
    data = {}
    try:
        xl = pd.ExcelFile(excel_file)
        for sheet_name in xl.sheet_names:
            data[sheet_name] = pd.read_excel(xl, sheet_name=sheet_name)
    except Exception as e:
        print(f"读取文件失败 {excel_file}: {e}")
    return data


def merge_data(all_dirs):
    """合并所有爬取数据"""
    # 存储所有数据，按区县和行业代码去重，保留最新数据
    all_data = {}  # key: (区县, 行业代码), value: 数据行
    
    print(f"\n找到 {len(all_dirs)} 个爬取结果目录:")
    for dir_path, dir_name in all_dirs:
        excel_file = os.path.join(dir_path, "数据统计.xlsx")
        print(f"  - {dir_name}")
        
        data = load_crawl_data(excel_file)
        if not data:
            continue
        
        # 处理每个区县的sheet
        for sheet_name, df in data.items():
            if sheet_name in ['总览', '区县汇总', '行业汇总']:
                continue
            
            # 跳过非区县的sheet
            if '行业代码' not in df.columns:
                continue
                
            for _, row in df.iterrows():
                district = sheet_name
                industry_code = row.get('行业代码', '')
                
                if not industry_code or industry_code == 'C':
                    continue
                
                key = (district, str(industry_code))
                # 保留最新的数据（后面遍历的会覆盖前面的）
                all_data[key] = {
                    '区县': district,
                    '行业代码': str(industry_code),
                    '行业名称': row.get('行业名称', ''),
                    '企业数量': row.get('企业数量', 0),
                    '来源目录': dir_name
                }
    
    return all_data


def create_merged_excel(all_data, output_file, districts, industries):
    """创建合并后的Excel文件"""
    # 创建工作簿
    wb = Workbook()
    
    # 删除默认Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 将数据转换为DataFrame
    data_list = list(all_data.values())
    df_all = pd.DataFrame(data_list)
    
    # 创建区县汇总数据
    district_summary = df_all.groupby('区县')['企业数量'].sum().reset_index()
    district_summary.columns = ['区县', '企业总数']
    
    # 创建行业汇总数据
    industry_summary = df_all.groupby(['行业代码', '行业名称'])['企业数量'].sum().reset_index()
    industry_summary.columns = ['行业代码', '行业名称', '企业总数']
    
    # 添加制造业合计行
    total_count = df_all['企业数量'].sum()
    
    # 创建总览表
    overview_sheet = wb.create_sheet(title="总览", index=0)
    overview_sheet['A1'] = "长沙市制造业企业统计（汇总）"
    overview_sheet['A1'].font = Font(bold=True, size=14)
    overview_sheet.merge_cells('A1:D1')
    overview_sheet['A2'] = f"汇总时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    overview_sheet['A2'].font = Font(italic=True)
    
    # 写入汇总信息
    overview_sheet['A4'] = "统计概览"
    overview_sheet['A4'].font = Font(bold=True)
    overview_sheet['A5'] = f"区县数量: {len(districts)}"
    overview_sheet['A6'] = f"行业数量: {len(industries)}"
    overview_sheet['A7'] = f"数据总条数: {len(data_list)}"
    overview_sheet['A8'] = f"企业总数: {total_count}"
    
    # 创建各区县工作表
    for district in districts:
        sheet = wb.create_sheet(title=district)
        
        # 写入表头
        headers = ['行业代码', '行业名称', '企业数量']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGNMENT
            cell.border = THIN_BORDER
        
        # 写入数据
        row_num = 2
        for industry_code, industry_name in industries.items():
            key = (district, industry_code)
            count = 0
            if key in all_data:
                count = all_data[key]['企业数量']
            
            sheet.cell(row=row_num, column=1, value=industry_code).border = THIN_BORDER
            sheet.cell(row=row_num, column=2, value=industry_name).border = THIN_BORDER
            sheet.cell(row=row_num, column=3, value=count).border = THIN_BORDER
            row_num += 1
        
        # 写入合计行
        total_cell = sheet.cell(row=row_num, column=1, value='C')
        total_cell.font = Font(bold=True)
        total_cell.border = THIN_BORDER
        sheet.cell(row=row_num, column=2, value='制造业合计').font = Font(bold=True)
        sheet.cell(row=row_num, column=2).border = THIN_BORDER
        
        # 计算该区县的合计
        district_total = sum(all_data.get((district, code), {}).get('企业数量', 0) 
                           for code in industries.keys())
        total_count_cell = sheet.cell(row=row_num, column=3, value=district_total)
        total_count_cell.font = Font(bold=True)
        total_count_cell.border = THIN_BORDER
        
        # 调整列宽
        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['B'].width = 35
        sheet.column_dimensions['C'].width = 12
    
    # 创建区县汇总表
    district_sheet = wb.create_sheet(title="区县汇总")
    headers = ['区县', '企业总数']
    for col, header in enumerate(headers, 1):
        cell = district_sheet.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    
    for row_num, (_, row) in enumerate(district_summary.iterrows(), 2):
        district_sheet.cell(row=row_num, column=1, value=row['区县']).border = THIN_BORDER
        district_sheet.cell(row=row_num, column=2, value=row['企业总数']).border = THIN_BORDER
    
    district_sheet.column_dimensions['A'].width = 15
    district_sheet.column_dimensions['B'].width = 12
    
    # 创建行业汇总表
    industry_sheet = wb.create_sheet(title="行业汇总")
    headers = ['行业代码', '行业名称', '企业总数']
    for col, header in enumerate(headers, 1):
        cell = industry_sheet.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    
    for row_num, (_, row) in enumerate(industry_summary.iterrows(), 2):
        industry_sheet.cell(row=row_num, column=1, value=row['行业代码']).border = THIN_BORDER
        industry_sheet.cell(row=row_num, column=2, value=row['行业名称']).border = THIN_BORDER
        industry_sheet.cell(row=row_num, column=3, value=row['企业总数']).border = THIN_BORDER
    
    # 添加合计行
    total_row = len(industry_summary) + 2
    industry_sheet.cell(row=total_row, column=1, value='C').font = Font(bold=True)
    industry_sheet.cell(row=total_row, column=2, value='制造业合计').font = Font(bold=True)
    industry_sheet.cell(row=total_row, column=3, value=total_count).font = Font(bold=True)
    
    industry_sheet.column_dimensions['A'].width = 12
    industry_sheet.column_dimensions['B'].width = 35
    industry_sheet.column_dimensions['C'].width = 12
    
    # 保存文件
    wb.save(output_file)
    print(f"\n汇总完成!")
    print(f"输出文件: {output_file}")
    print(f"  - 区县数: {len(districts)}")
    print(f"  - 行业数: {len(industries)}")
    print(f"  - 企业总数: {total_count}")


def main():
    print("=" * 60)
    print("数据汇总工具 - 合并多次爬取结果")
    print("=" * 60)
    
    # 获取所有爬取目录
    all_dirs = get_all_crawl_dirs()
    
    if not all_dirs:
        print("\n未找到任何爬取结果目录")
        return
    
    # 合并数据
    all_data = merge_data(all_dirs)
    
    if not all_data:
        print("\n没有有效数据可合并")
        return
    
    # 创建合并后的Excel
    output_file = config.MERGED_OUTPUT_FILE
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    create_merged_excel(
        all_data, 
        output_file, 
        config.CHANGSHA_DISTRICTS, 
        config.MANUFACTURING_SUBCATEGORIES
    )
    
    # 验证数据一致性
    print("\n数据验证:")
    df_all = pd.DataFrame(list(all_data.values()))
    
    for district in config.CHANGSHA_DISTRICTS:
        district_data = df_all[df_all['区县'] == district]
        sub_sum = district_data['企业数量'].sum()
        print(f"  {district}: {sub_sum} 家企业")


if __name__ == "__main__":
    main()
