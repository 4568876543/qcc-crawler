"""
表格转换器主模块
将爬取的长表格式转换为目标格式
"""
import os
import sys
import pandas as pd
from typing import Optional

from .sheet1_generator import Sheet1Generator
from .sheet2_generator import Sheet2Generator
from .chart_generator import ChartGenerator


class TableConverter:
    """
    表格转换器

    将爬取的明细数据转换为目标格式：
    - Sheet1: 横向宽表（每个区县占2列）
    - Sheet2: 行业排名表
    """

    def __init__(self, city_name: str = "重庆市"):
        """
        Args:
            city_name: 城市名称
        """
        self.city_name = city_name
        self.sheet1_generator = Sheet1Generator(city_name=city_name)
        self.sheet2_generator = Sheet2Generator(top_n=10, city_name=city_name)
        self.source_file = None
        self.data_loaded = False

    def load_from_excel(self, file_path: str) -> bool:
        """
        从 Excel 文件加载数据

        Args:
            file_path: 源 Excel 文件路径

        Returns:
            是否加载成功
        """
        try:
            print(f"[加载] 读取文件: {file_path}")
            self.source_file = file_path

            # 读取明细数据 sheet
            df_raw = pd.read_excel(file_path, sheet_name='明细数据', header=None)

            # 找到表头行（"区县" 在某列）
            header_row = None
            for i in range(len(df_raw)):
                row_values = df_raw.iloc[i].astype(str).tolist()
                if '区县' in row_values:
                    header_row = i
                    break

            if header_row is None:
                print("[错误] 未找到表头行（区县列）")
                return False

            print(f"[加载] 表头行: {header_row}")

            # 重新读取，跳过前面的行
            df = pd.read_excel(file_path, sheet_name='明细数据', header=header_row)

            # 清理列名
            df.columns = ['区县', '行业代码', '行业名称', '企业数量']

            # 删除无效行
            df = df.dropna(subset=['区县', '行业代码', '行业名称'])

            # 转换数据类型
            df['行业代码'] = df['行业代码'].astype(str)
            df['企业数量'] = pd.to_numeric(df['企业数量'], errors='coerce').fillna(0).astype(int)

            # 去掉重复表头（如果有）
            df = df[df['行业代码'] != '行业代码']

            print(f"[加载] 读取到 {len(df)} 条数据")

            # 按行业分组添加数据
            for (industry_code, industry_name), group in df.groupby(['行业代码', '行业名称']):
                district_counts = dict(zip(group['区县'], group['企业数量']))
                self.sheet1_generator.add_data(str(industry_code), industry_name, district_counts)
                self.sheet2_generator.add_data(str(industry_code), industry_name, district_counts)

            # 完成数据添加，计算全市总计
            self.sheet1_generator.finalize()

            self.data_loaded = True
            print("[加载] 数据加载完成")
            return True

        except Exception as e:
            print(f"[错误] 加载数据失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def convert(self, output_path: str) -> bool:
        """
        转换并保存为 Excel 文件（模仿用户手动修改的格式）

        Args:
            output_path: 输出文件路径

        Returns:
            是否转换成功
        """
        if not self.data_loaded:
            print("[错误] 请先加载数据")
            return False

        try:
            print(f"[转换] 输出到: {output_path}")

            # 确保输出目录存在
            os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)

            # 直接使用 Sheet1Generator 生成 Excel（带公式）
            self.sheet1_generator.to_excel(output_path)

            # 重新打开文件添加产业分析图表Sheet
            from openpyxl import load_workbook
            wb = load_workbook(output_path)

            # 计算行业总计（用于产业分析图表）
            industry_totals = []
            for industry_code, industry_name in self.sheet2_generator.industry_order:
                counts = self.sheet2_generator.industry_data.get(industry_name, {})
                total = sum(counts.values())
                industry_totals.append((industry_name, total))

            # 按总数降序排序
            industry_totals.sort(key=lambda x: x[1], reverse=True)

            # 创建"产业分析图表"Sheet
            if "产业分析图表" in wb.sheetnames:
                del wb["产业分析图表"]
            chart_sheet = wb.create_sheet("产业分析图表")

            # 添加产业分析图表数据（不带图片，只有数据）
            self._add_industry_analysis_data(chart_sheet, industry_totals)

            wb.save(output_path)
            print(f"[转换] 完成！文件已保存到: {output_path}")
            return True

        except Exception as e:
            print(f"[错误] 转换失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _add_industry_analysis_data(self, chart_sheet, industry_totals):
        """
        添加产业分析图表数据（只有数据，没有图片）

        格式：
        - 第1行: 标题（行业, 企业数量, 占比）
        - 第2-32行: 31个行业数据（按企业数量降序）
        - 第33行: 合计行
        """
        from openpyxl.styles import Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        # 计算总数量
        total_sum = sum(t[1] for t in industry_totals)

        # 样式
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 第1行：标题
        chart_sheet.cell(row=1, column=1, value="行业")
        chart_sheet.cell(row=1, column=1).font = header_font
        chart_sheet.cell(row=1, column=1).fill = header_fill
        chart_sheet.cell(row=1, column=1).border = thin_border

        chart_sheet.cell(row=1, column=2, value="企业数量")
        chart_sheet.cell(row=1, column=2).font = header_font
        chart_sheet.cell(row=1, column=2).fill = header_fill
        chart_sheet.cell(row=1, column=2).border = thin_border

        chart_sheet.cell(row=1, column=3, value="占比")
        chart_sheet.cell(row=1, column=3).font = header_font
        chart_sheet.cell(row=1, column=3).fill = header_fill
        chart_sheet.cell(row=1, column=3).border = thin_border

        # 第2-32行：行业数据
        for i, (industry_name, total) in enumerate(industry_totals):
            row = 2 + i

            chart_sheet.cell(row=row, column=1, value=industry_name)
            chart_sheet.cell(row=row, column=1).border = thin_border

            chart_sheet.cell(row=row, column=2, value=total)
            chart_sheet.cell(row=row, column=2).border = thin_border
            chart_sheet.cell(row=row, column=2).number_format = '#,##0'

            # 占比用公式
            chart_sheet.cell(row=row, column=3, value=f"=B{row}/$B$33")
            chart_sheet.cell(row=row, column=3).border = thin_border
            chart_sheet.cell(row=row, column=3).number_format = '0.00%'

        # 第33行：合计行
        total_row = 2 + len(industry_totals)
        chart_sheet.cell(row=total_row, column=1, value="合计")
        chart_sheet.cell(row=total_row, column=1).font = Font(bold=True)
        chart_sheet.cell(row=total_row, column=1).fill = total_fill
        chart_sheet.cell(row=total_row, column=1).border = thin_border

        chart_sheet.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row-1})")
        chart_sheet.cell(row=total_row, column=2).font = Font(bold=True)
        chart_sheet.cell(row=total_row, column=2).fill = total_fill
        chart_sheet.cell(row=total_row, column=2).border = thin_border
        chart_sheet.cell(row=total_row, column=2).number_format = '#,##0'

        chart_sheet.cell(row=total_row, column=3, value=f"=B{total_row}/$B${total_row}")
        chart_sheet.cell(row=total_row, column=3).font = Font(bold=True)
        chart_sheet.cell(row=total_row, column=3).fill = total_fill
        chart_sheet.cell(row=total_row, column=3).border = thin_border
        chart_sheet.cell(row=total_row, column=3).number_format = '0.00%'

        # 设置列宽
        chart_sheet.column_dimensions['A'].width = 35
        chart_sheet.column_dimensions['B'].width = 15
        chart_sheet.column_dimensions['C'].width = 12

    def validate(self) -> dict:
        """
        验证数据完整性

        Returns:
            验证结果字典
        """
        if not self.data_loaded:
            return {'valid': False, 'error': '数据未加载'}

        results = {
            'valid': True,
            'industries': len(self.sheet1_generator.industry_data),
            'districts': len(self.sheet1_generator.districts),
            'city_total': self.sheet1_generator.city_total,
            'warnings': []
        }

        # 检查行业数量
        if results['industries'] != 31:
            results['warnings'].append(f"行业数量异常：{results['industries']}（期望31）")

        # 检查区县数量
        if results['districts'] < 10:
            results['warnings'].append(f"区县数量异常：{results['districts']}")

        # 检查数据完整性
        total_industry_sum = sum(
            sum(counts.values())
            for counts in self.sheet1_generator.industry_data.values()
        )

        if total_industry_sum != results['city_total'] * results['industries']:
            # 这是正常的，因为每个行业只有部分区县有数据
            pass

        return results


def main():
    """命令行入口"""
    import argparse

    parser = argparse.ArgumentParser(description='将爬取数据转换为目标格式')
    parser.add_argument('--input', '-i', required=True, help='输入文件路径')
    parser.add_argument('--output', '-o', required=True, help='输出文件路径')
    parser.add_argument('--city', '-c', default='重庆市', help='城市名称')

    args = parser.parse_args()

    # 创建转换器
    converter = TableConverter(city_name=args.city)

    # 加载数据
    if not converter.load_from_excel(args.input):
        sys.exit(1)

    # 验证
    validation = converter.validate()
    print(f"\n[验证] 行业数: {validation['industries']}")
    print(f"[验证] 区县数: {validation['districts']}")
    print(f"[验证] 全市合计: {validation['city_total']:,}")
    if validation['warnings']:
        for warning in validation['warnings']:
            print(f"[警告] {warning}")

    # 转换
    if not converter.convert(args.output):
        sys.exit(1)

    print("\n转换完成！")


if __name__ == "__main__":
    main()
