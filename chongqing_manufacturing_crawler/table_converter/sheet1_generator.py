"""
Sheet1 生成器 - 生成横向宽表格式（模仿用户手动修改的格式）
每个区县占2列：[数量] [占比]
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


class Sheet1Generator:
    """生成横向宽表（用户手动修改的格式）"""

    def __init__(self, city_name="重庆市"):
        self.city_name = city_name
        self.industry_data = {}  # {行业名称: {区县: 数量}}
        self.industry_order = []  # 保持行业顺序
        self.districts = []  # 区县列表
        self.total_by_district = {}  # 各区县合计
        self.city_total = 0  # 全市总计

    def add_data(self, industry_code, industry_name, district_counts):
        """
        添加行业数据
        """
        # 排除制造业合计行（如果数据中有）
        if industry_code == 'C' or industry_name == '制造业合计':
            for district, count in district_counts.items():
                if district not in self.total_by_district:
                    self.total_by_district[district] = 0
                self.total_by_district[district] += count
            return

        self.industry_data[industry_name] = district_counts
        self.industry_order.append((industry_code, industry_name))

        # 累加区县合计
        for district, count in district_counts.items():
            if district not in self.total_by_district:
                self.total_by_district[district] = 0
            self.total_by_district[district] += count

        # 更新区县列表
        for district in district_counts.keys():
            if district not in self.districts:
                self.districts.append(district)

    def finalize(self):
        """完成数据添加后调用，计算全市总计"""
        self.city_total = sum(self.total_by_district.values())

    def to_excel(self, output_path: str):
        """
        直接输出为 Excel 文件（模仿用户手动修改的格式）

        格式：
        - 第1行: 标题行（行业分类, 长沙市, 占比, 天心区, 占比, ...）
        - 第2-32行: 31个行业数据
        - 第33行: 合计行
        - 第34行: 验证行
        """
        # 按行业代码排序
        self.industry_order.sort(key=lambda x: int(str(x[0])) if str(x[0]).isdigit() else 0)
        self.districts.sort()

        wb = Workbook()
        ws = wb.active
        ws.title = "明细"

        # 样式定义
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')

        # ========== 第1行：标题行 ==========
        # 列结构：行业分类（31大类）, 长沙市, 占比, 天心区, 占比, 宁乡市, 占比, ...
        # A列=行业分类, B列=长沙市数量, C列=长沙市占比, D列=天心区数量, E列=天心区占比, ...

        ws.cell(row=1, column=1, value="行业分类（31大类）")
        ws.cell(row=1, column=1).font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).border = thin_border
        ws.cell(row=1, column=1).alignment = center_align

        ws.cell(row=1, column=2, value=self.city_name)
        ws.cell(row=1, column=2).font = header_font
        ws.cell(row=1, column=2).fill = header_fill
        ws.cell(row=1, column=2).border = thin_border
        ws.cell(row=1, column=2).alignment = center_align

        ws.cell(row=1, column=3, value="占比")
        ws.cell(row=1, column=3).font = header_font
        ws.cell(row=1, column=3).fill = header_fill
        ws.cell(row=1, column=3).border = thin_border
        ws.cell(row=1, column=3).alignment = center_align

        col_idx = 4
        for district in self.districts:
            ws.cell(row=1, column=col_idx, value=district)
            ws.cell(row=1, column=col_idx).font = header_font
            ws.cell(row=1, column=col_idx).fill = header_fill
            ws.cell(row=1, column=col_idx).border = thin_border
            ws.cell(row=1, column=col_idx).alignment = center_align

            ws.cell(row=1, column=col_idx + 1, value="占比")
            ws.cell(row=1, column=col_idx + 1).font = header_font
            ws.cell(row=1, column=col_idx + 1).fill = header_fill
            ws.cell(row=1, column=col_idx + 1).border = thin_border
            ws.cell(row=1, column=col_idx + 1).alignment = center_align

            col_idx += 2

        # ========== 第2-32行：行业数据 ==========
        # B列 = 城市数量（列号2）
        # C列 = 城市占比（列号3）
        # D,F,H... = 区县数量
        # E,G,I... = 区县占比

        data_start_row = 2
        data_end_row = 2 + len(self.industry_order) - 1  # 第32行
        total_row = data_end_row + 1  # 第33行

        for idx, (industry_code, industry_name) in enumerate(self.industry_order):
            excel_row = data_start_row + idx
            counts = self.industry_data[industry_name]

            # A列：行业名称
            ws.cell(row=excel_row, column=1, value=industry_name)
            ws.cell(row=excel_row, column=1).font = Font(bold=True)
            ws.cell(row=excel_row, column=1).border = thin_border

            # B列：城市数量（数值）
            city_sum = sum(counts.values())
            ws.cell(row=excel_row, column=2, value=city_sum)
            ws.cell(row=excel_row, column=2).border = thin_border
            ws.cell(row=excel_row, column=2).number_format = '#,##0'

            # C列：城市占比（公式 =B2/$B$33）
            col_letter = get_column_letter
            ws.cell(row=excel_row, column=3, value=f"=B{excel_row}/$B${total_row}")
            ws.cell(row=excel_row, column=3).border = thin_border
            ws.cell(row=excel_row, column=3).number_format = '0.00%'

            # 各区县数量和占比
            col_idx = 4
            for district in self.districts:
                count = counts.get(district, 0)

                # 数量列（数值）
                ws.cell(row=excel_row, column=col_idx, value=count)
                ws.cell(row=excel_row, column=col_idx).border = thin_border
                ws.cell(row=excel_row, column=col_idx).number_format = '#,##0'

                # 占比列（公式 =D2/$D$33）
                district_col_letter = get_column_letter(col_idx)
                ws.cell(row=excel_row, column=col_idx + 1,
                       value=f"={district_col_letter}{excel_row}/${district_col_letter}${total_row}")
                ws.cell(row=excel_row, column=col_idx + 1).border = thin_border
                ws.cell(row=excel_row, column=col_idx + 1).number_format = '0.00%'

                col_idx += 2

        # ========== 第33行：合计行 ==========
        ws.cell(row=total_row, column=1, value="合计")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=1).fill = total_fill
        ws.cell(row=total_row, column=1).border = thin_border

        # B列合计：=SUM(B2:B32)
        ws.cell(row=total_row, column=2, value=f"=SUM(B{data_start_row}:B{data_end_row})")
        ws.cell(row=total_row, column=2).font = Font(bold=True)
        ws.cell(row=total_row, column=2).fill = total_fill
        ws.cell(row=total_row, column=2).border = thin_border
        ws.cell(row=total_row, column=2).number_format = '#,##0'

        # C列占比：=B33/$B$33 (即100%)
        ws.cell(row=total_row, column=3, value=f"=B{total_row}/$B${total_row}")
        ws.cell(row=total_row, column=3).font = Font(bold=True)
        ws.cell(row=total_row, column=3).fill = total_fill
        ws.cell(row=total_row, column=3).border = thin_border
        ws.cell(row=total_row, column=3).number_format = '0.00%'

        # 各区县合计
        col_idx = 4
        for district in self.districts:
            district_col_letter = get_column_letter(col_idx)
            district_col_letter_end = get_column_letter(col_idx + 1)

            # 数量列：=SUM(D2:D32)
            ws.cell(row=total_row, column=col_idx,
                   value=f"=SUM({district_col_letter}{data_start_row}:{district_col_letter}{data_end_row})")
            ws.cell(row=total_row, column=col_idx).font = Font(bold=True)
            ws.cell(row=total_row, column=col_idx).fill = total_fill
            ws.cell(row=total_row, column=col_idx).border = thin_border
            ws.cell(row=total_row, column=col_idx).number_format = '#,##0'

            # 占比列：=D33/$D$33
            ws.cell(row=total_row, column=col_idx + 1,
                   value=f"={district_col_letter}{total_row}/${district_col_letter}${total_row}")
            ws.cell(row=total_row, column=col_idx + 1).font = Font(bold=True)
            ws.cell(row=total_row, column=col_idx + 1).fill = total_fill
            ws.cell(row=total_row, column=col_idx + 1).border = thin_border
            ws.cell(row=total_row, column=col_idx + 1).number_format = '0.00%'

            col_idx += 2

        # ========== 第34行：验证行 ==========
        # B34 = D33+F33+H33+J33+L33+N33+P33+R33+T33+V33
        verify_row = total_row + 1
        ws.cell(row=verify_row, column=2, value=self._generate_verify_formula(total_row))
        ws.cell(row=verify_row, column=2).font = Font(bold=True)

        # 设置列宽
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        for i, district in enumerate(self.districts):
            col_letter = get_column_letter(4 + i * 2)  # D, F, H, ...
            ws.column_dimensions[col_letter].width = 12
            ws.column_dimensions[get_column_letter(4 + i * 2 + 1)].width = 8

        wb.save(output_path)
        print(f"[Sheet1] 已保存到: {output_path}")
        return output_path

    def _generate_verify_formula(self, total_row: int) -> str:
        """生成验证公式：B34=D33+F33+H33+J33+L33+N33+P33+R33+T33+V33"""
        cols = []
        col_idx = 4  # 从D列开始（第一个区县）
        for _ in self.districts:
            col_letter = get_column_letter(col_idx)
            cols.append(f"{col_letter}{total_row}")
            col_idx += 2
        return f"=B{total_row}=" + "+".join(cols)


if __name__ == "__main__":
    # 测试代码
    generator = Sheet1Generator("长沙市")

    # 模拟数据
    generator.add_data('C', '制造业合计', {
        '天心区': 5000, '宁乡市': 10000, '岳麓区': 8000
    })

    generator.add_data('13', '农副食品加工业', {
        '天心区': 436, '宁乡市': 1648, '岳麓区': 720
    })

    generator.add_data('14', '食品制造业', {
        '天心区': 798, '宁乡市': 1065, '岳麓区': 1600
    })

    generator.finalize()
    generator.to_excel('/tmp/test_output.xlsx')
    print("测试完成")
