"""
图表生成器 - 通用模块
使用matplotlib生成高质量图表，插入到Excel
"""
import os
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import numpy as np
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
import io


# 设置中文字体
plt.rcParams["font.sans-serif"] = ["SimHei", "WenQuanYi Zen Hei", "Arial Unicode MS"]
plt.rcParams["axes.unicode_minus"] = False

# 默认数据文件夹
DEFAULT_DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
DEFAULT_OUTPUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "output")


class ChartGenerator:
    """通用图表生成器"""

    def __init__(self, city_name: str = "重庆市"):
        """
        Args:
            city_name: 城市名称
        """
        self.city_name = city_name

    def generate_all_charts(self, industry_totals: list, industry_data: dict) -> dict:
        """
        生成所有图表

        Args:
            industry_totals: list of (行业名称, 企业数量)，已按降序排列
            industry_data: dict {行业名称: {区县: 数量}}

        Returns:
            dict: {图表名称: BytesIO对象}
        """
        # 计算总数量
        total_sum = sum(t[1] for t in industry_totals)

        charts = {}

        # 图表1：环形饼图
        charts['pie'] = self._generate_pie_chart(industry_totals, total_sum)

        # 图表2：横向条形图
        charts['bar'] = self._generate_bar_chart(industry_totals, total_sum)

        # 图表3：柱状图+折线组合图
        charts['pareto'] = self._generate_pareto_chart(industry_data, total_sum)

        return charts

    def _generate_pie_chart(self, industry_totals: list, total_sum: float) -> io.BytesIO:
        """生成环形饼图"""
        # 分离前10和"其他"
        top_10 = industry_totals[:10]
        other_sum = sum(t[1] for t in industry_totals[10:])
        other_ratio = other_sum / total_sum if total_sum > 0 else 0

        pie_labels = [t[0] for t in top_10] + (["其他"] if other_sum > 0 else [])
        pie_values = [t[1] for t in top_10] + ([other_sum] if other_sum > 0 else [])
        pie_percentages = [t[1]/total_sum*100 for t in top_10] + ([other_ratio*100] if other_sum > 0 else [])

        colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7',
                  '#DDA0DD', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E9', '#B0B0B0']

        fig, ax = plt.subplots(figsize=(12, 10))
        wedges, texts, autotexts = ax.pie(
            pie_values,
            labels=None,
            autopct='%1.1f%%',
            startangle=90,
            colors=colors[:len(pie_values)],
            wedgeprops={'width': 0.55, 'edgecolor': 'white', 'linewidth': 2},
            pctdistance=0.7
        )

        for autotext in autotexts:
            autotext.set_fontsize(9)
            autotext.set_fontweight('bold')
            autotext.set_color('white')

        ax.text(0, 0, f'{self.city_name}\n制造业', ha='center', va='center',
                fontsize=18, fontweight='bold', color='#333333')

        legend_labels = [f'{name} ({pct:.1f}%)' for name, pct in zip(pie_labels, pie_percentages)]
        ax.legend(wedges, legend_labels, title="行业占比", loc="center left",
                 bbox_to_anchor=(0.95, 0, 0.5, 1), fontsize=8, title_fontsize=9)

        plt.title(f"{self.city_name}制造业31大类占比分布", fontsize=14, fontweight='bold', pad=10)
        plt.tight_layout()

        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=120, bbox_inches='tight',
                   facecolor='white', edgecolor='none')
        buffer.seek(0)
        plt.close(fig)

        return buffer

    def _generate_bar_chart(self, industry_totals: list, total_sum: float) -> io.BytesIO:
        """生成横向条形图（Top 10）"""
        bar_data = industry_totals[:10]
        bar_names = [t[0] for t in bar_data]
        bar_values = [t[1] for t in bar_data]

        fig, ax = plt.subplots(figsize=(12, 9))
        y_pos = np.arange(len(bar_names))
        bars = ax.barh(y_pos, bar_values, color='#4ECDC4', edgecolor='white', linewidth=0.5, height=0.7)

        for bar, val in zip(bars, bar_values):
            ax.text(val + bar_values[0]*0.02, bar.get_y() + bar.get_height()/2,
                   f'{val:,} ({val/total_sum*100:.1f}%)',
                   va='center', ha='left', fontsize=9, color='#333333')

        ax.set_yticks(y_pos)
        ax.set_yticklabels(bar_names, fontsize=9)
        ax.set_xlabel('企业数量', fontsize=10)
        ax.set_xlim(0, max(bar_values) * 1.4)
        ax.invert_yaxis()
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)

        plt.title(f"{self.city_name}前10大制造业类别规模排名", fontsize=14, fontweight='bold', pad=10)
        plt.tight_layout()

        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=120, bbox_inches='tight',
                   facecolor='white', edgecolor='none')
        buffer.seek(0)
        plt.close(fig)

        return buffer

    def _generate_pareto_chart(self, industry_data: dict, total_sum: float) -> io.BytesIO:
        """生成柱状图+折线图组合（帕累托图）"""
        # 计算各区县合计
        district_totals = {}
        for industry_name, counts in industry_data.items():
            for district, count in counts.items():
                if district not in district_totals:
                    district_totals[district] = 0
                district_totals[district] += count

        # 按数量降序排序
        sorted_districts = sorted(district_totals.items(), key=lambda x: x[1], reverse=True)
        district_names = [d[0] for d in sorted_districts]
        district_values = [d[1] for d in sorted_districts]

        # 计算累积百分比
        cumsum = np.cumsum(district_values)
        cumsum_percent = cumsum / total_sum * 100

        fig, ax1 = plt.subplots(figsize=(12, 9))

        # 柱状图
        x_pos = np.arange(len(district_names))
        bars = ax1.bar(x_pos, district_values, color='#4ECDC4', edgecolor='white',
                       linewidth=0.5, width=0.6, label='企业数量')

        for bar, val in zip(bars, district_values):
            ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + total_sum*0.01,
                    f'{val:,}', ha='center', va='bottom', fontsize=7, color='#333333')

        # 折线图（累积百分比）
        ax2 = ax1.twinx()
        ax2.plot(x_pos, cumsum_percent, color='#FF6B6B', marker='o',
                linewidth=2, markersize=4, label='累积占比')

        # 80%参考线
        ax2.axhline(y=80, color='gray', linestyle='--', linewidth=1, alpha=0.7)
        ax2.text(len(district_names)-1, 82, '80%', fontsize=8, color='gray')

        ax1.set_xlabel('区县', fontsize=10)
        ax1.set_ylabel('企业数量', fontsize=10, color='#4ECDC4')
        ax2.set_ylabel('累积占比 (%)', fontsize=10, color='#FF6B6B')
        ax2.set_ylim(0, 110)

        ax1.set_xticks(x_pos)
        ax1.set_xticklabels(district_names, rotation=45, ha='right', fontsize=8)
        ax1.tick_params(axis='y', labelcolor='#4ECDC4')
        ax2.tick_params(axis='y', labelcolor='#FF6B6F')

        ax1.spines['top'].set_visible(False)
        ax2.spines['top'].set_visible(False)

        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper right', fontsize=8)

        plt.title(f"{self.city_name}各区县企业数量分布及累积占比", fontsize=14, fontweight='bold', pad=10)
        plt.tight_layout()

        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=120, bbox_inches='tight',
                   facecolor='white', edgecolor='none')
        buffer.seek(0)
        plt.close(fig)

        return buffer

    def add_charts_to_sheet(self, chart_sheet, industry_totals: list, industry_data: dict):
        """
        将图表添加到指定sheet

        Args:
            chart_sheet: openpyxl Worksheet对象
            industry_totals: list of (行业名称, 企业数量)
            industry_data: dict {行业名称: {区县: 数量}}
        """
        total_sum = sum(t[1] for t in industry_totals)

        # 写入数据（A-C列）
        chart_sheet.cell(row=1, column=1, value="行业").font = Font(bold=True)
        chart_sheet.cell(row=1, column=2, value="企业数量").font = Font(bold=True)
        chart_sheet.cell(row=1, column=3, value="占比").font = Font(bold=True)

        for i, (industry_name, total) in enumerate(industry_totals):
            row = 2 + i
            chart_sheet.cell(row=row, column=1, value=industry_name)
            chart_sheet.cell(row=row, column=2, value=total).number_format = '#,##0'
            ratio = total / total_sum if total_sum > 0 else 0
            chart_sheet.cell(row=row, column=3, value=ratio).number_format = '0.00%'

        # 生成图表
        charts = self.generate_all_charts(industry_totals, industry_data)

        # 插入图表（从F列开始，三个图从上到下排列）
        chart_start_col = 'F'

        img_pie = XLImage(charts['pie'])
        img_pie.anchor = f'{chart_start_col}1'
        chart_sheet.add_image(img_pie)

        img_bar = XLImage(charts['bar'])
        img_bar.anchor = f'{chart_start_col}22'
        chart_sheet.add_image(img_bar)

        img_pareto = XLImage(charts['pareto'])
        img_pareto.anchor = f'{chart_start_col}43'
        chart_sheet.add_image(img_pareto)

        # 设置列宽
        chart_sheet.column_dimensions['A'].width = 35
        chart_sheet.column_dimensions['B'].width = 15
        chart_sheet.column_dimensions['C'].width = 12


def generate_charts_image(industry_totals: list, industry_data: dict,
                          city_name: str = "重庆市",
                          output_path: str = None) -> dict:
    """
    快捷函数：生成所有图表并可选保存

    Args:
        industry_totals: list of (行业名称, 企业数量)
        industry_data: dict {行业名称: {区县: 数量}}
        city_name: 城市名称
        output_path: 可选，保存路径

    Returns:
        dict: {图表名称: BytesIO对象}
    """
    generator = ChartGenerator(city_name=city_name)
    charts = generator.generate_all_charts(industry_totals, industry_data)

    if output_path:
        for name, buffer in charts.items():
            with open(f"{output_path}_{name}.png", 'wb') as f:
                f.write(buffer.getvalue())

    return charts
