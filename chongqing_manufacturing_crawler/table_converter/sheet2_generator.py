"""
Sheet2 生成器 - 生成横向行业排名表
每个行业占2列：[地区] [数值]
每个行业取前N名区县（默认10名）
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList


class Sheet2Generator:
    """生成 Sheet2 格式的横向排名表"""

    def __init__(self, top_n: int = 10, city_name: str = "重庆市"):
        """
        Args:
            top_n: 每个行业取前几名区县
            city_name: 城市名称
        """
        self.top_n = top_n
        self.city_name = city_name
        self.industry_data = {}  # {行业名称: {区县: 数量}} 按数量降序
        self.industry_order = []  # 保持行业顺序
        self.districts = []  # 区县列表（用于排序）

    def add_data(self, industry_code, industry_name, district_counts):
        """
        添加行业数据（会自动排序取前N）

        Args:
            industry_code: 行业代码
            industry_name: 行业名称
            district_counts: dict {区县: 企业数量}
        """
        # 排除制造业合计行
        if industry_code == 'C' or industry_name == '制造业合计':
            return

        # 按数量降序排序，取前N
        sorted_districts = sorted(district_counts.items(), key=lambda x: x[1], reverse=True)
        top_districts = dict(sorted_districts[:self.top_n])

        self.industry_data[industry_name] = top_districts
        self.industry_order.append((industry_code, industry_name))

        # 更新区县列表（按全市数量排序的前10）
        for district in top_districts.keys():
            if district not in self.districts:
                self.districts.append(district)

        # 按行业代码排序
        self.industry_order.sort(key=lambda x: int(str(x[0])) if str(x[0]).isdigit() else 0)

    def generate(self) -> pd.DataFrame:
        """
        生成 DataFrame

        Returns:
            DataFrame 每个行业2列（地区 + 数值）
        """
        if not self.industry_data:
            raise ValueError("没有添加任何行业数据")

        # 获取所有区县（按第一个行业排序）
        all_districts = []
        for industry_name in [name for _, name in self.industry_order]:
            for district in self.industry_data[industry_name].keys():
                if district not in all_districts:
                    all_districts.append(district)

        # 构建数据
        data = []
        for district in all_districts:
            row = {}
            for industry_code, industry_name in self.industry_order:
                counts = self.industry_data[industry_name]
                if district in counts:
                    row[f'地区（{industry_name}）'] = district
                    row[f'数值'] = counts[district]
                else:
                    row[f'地区（{industry_name}）'] = district
                    row[f'数值'] = 0
            data.append(row)

        # 创建 DataFrame
        columns = []
        for industry_code, industry_name in self.industry_order:
            columns.append(f'地区（{industry_name}）')
            columns.append('数值')

        df = pd.DataFrame(data, columns=columns)

        return df

    def to_excel(self, output_path: str):
        """
        直接输出为 Excel 文件

        Args:
            output_path: 输出文件路径
        """
        df = self.generate()

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet2"

        # 样式定义
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')

        # 第一行：标题
        ws.cell(row=1, column=1, value=f"{self.city_name}制造业各行业企业数量排名")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        # 第二行：行业名称（每2列一组）
        col_idx = 1
        industry_names = [name for _, name in self.industry_order]
        for i, industry_name in enumerate(industry_names):
            ws.cell(row=2, column=col_idx, value=f'地区（{industry_name}）')
            ws.cell(row=2, column=col_idx).font = header_font
            ws.cell(row=2, column=col_idx).fill = header_fill
            ws.cell(row=2, column=col_idx).border = thin_border
            ws.cell(row=2, column=col_idx).alignment = center_align

            ws.cell(row=2, column=col_idx + 1, value=f'数值')
            ws.cell(row=2, column=col_idx + 1).font = header_font
            ws.cell(row=2, column=col_idx + 1).fill = header_fill
            ws.cell(row=2, column=col_idx + 1).border = thin_border
            ws.cell(row=2, column=col_idx + 1).alignment = center_align

            col_idx += 2

        # 数据行
        for row_idx, row_data in df.iterrows():
            excel_row = row_idx + 3

            col_idx = 1
            for col_name in df.columns:
                value = row_data[col_name]
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.border = thin_border

                if '数值' in col_name and isinstance(value, (int, float)):
                    cell.number_format = '#,##0'

        # 设置列宽
        col_idx = 1
        for industry_code, industry_name in self.industry_order:
            ws.column_dimensions[self._get_column_letter(col_idx)].width = 20
            ws.column_dimensions[self._get_column_letter(col_idx + 1)].width = 10
            col_idx += 2

        # 添加图表数据区域和图表
        self._add_charts(wb, ws)

        wb.save(output_path)
        print(f"[Sheet2] 已保存到: {output_path}")

        return output_path

    def _add_charts(self, wb, ws):
        """添加饼图和条形图"""
        # 计算每个行业的全市总计
        industry_totals = []
        for industry_code, industry_name in self.industry_order:
            counts = self.industry_data.get(industry_name, {})
            total = sum(counts.values())
            industry_totals.append((industry_name, total))

        # 计算总规模
        total_sum = sum(t[1] for t in industry_totals)

        # 图表数据起始行（Sheet2数据结束后空2行）
        chart_data_start_row = len(self.industry_order) + 5

        # 写入图表数据：行业名称、总规模、占比
        ws.cell(row=chart_data_start_row, column=1, value="行业").font = Font(bold=True)
        ws.cell(row=chart_data_start_row, column=2, value="企业数量").font = Font(bold=True)
        ws.cell(row=chart_data_start_row, column=3, value="占比").font = Font(bold=True)

        for i, (industry_name, total) in enumerate(industry_totals):
            row = chart_data_start_row + 1 + i
            ws.cell(row=row, column=1, value=industry_name)
            ws.cell(row=row, column=2, value=total).number_format = '#,##0'
            ratio = total / total_sum if total_sum > 0 else 0
            ws.cell(row=row, column=3, value=ratio).number_format = '0.00%'

        # 获取图表数据范围
        data_end_row = chart_data_start_row + len(industry_totals)
        data_ref = f"A{chart_data_start_row}:C{data_end_row}"

        # 创建"产业分析图表"Sheet
        if "产业分析图表" in wb.sheetnames:
            del wb["产业分析图表"]
        chart_sheet = wb.create_sheet("产业分析图表")

        # ===== 饼图：行业占比分布 =====
        pie = PieChart()
        pie.title = f"{self.city_name}制造业31大类占比分布"
        pie.dataLabels = DataLabelList()
        pie.dataLabels.showPercent = True
        pie.dataLabels.showCatName = True
        pie.dataLabels.showVal = False

        # 饼图数据：行业名称（列A）和占比（列C）
        labels = Reference(chart_sheet, min_col=1, min_row=chart_data_start_row + 1, max_row=data_end_row)
        pie_data = Reference(chart_sheet, min_col=3, min_row=chart_data_start_row, max_row=data_end_row)
        pie.add_data(pie_data, titles_from_data=True)
        pie.set_categories(labels)
        pie.width = 15
        pie.height = 10

        # 将饼图放到产业分析图表 sheet
        chart_sheet.add_chart(pie, "D2")

        # ===== 条形图：前15大行业排名 =====
        bar = BarChart()
        bar.type = "bar"  # 横向条形图
        bar.title = f"{self.city_name}前{self.top_n}大制造业类别规模排名"
        bar.y_axis.title = "行业"
        bar.x_axis.title = "企业数量"
        bar.x_axis.numFmt = '#,##0'

        # 取前15个行业（或top_n）
        bar_data_end_row = chart_data_start_row + min(15, len(industry_totals))
        bar_data_ref = f"A{chart_data_start_row}:B{bar_data_end_row}"

        # 条形图数据：行业名称（列A）和数量（列B）
        bar_labels = Reference(chart_sheet, min_col=1, min_row=chart_data_start_row + 1, max_row=bar_data_end_row)
        bar_values = Reference(chart_sheet, min_col=2, min_row=chart_data_start_row, max_row=bar_data_end_row)
        bar.add_data(bar_values, titles_from_data=True)
        bar.set_categories(bar_labels)
        bar.width = 18
        bar.height = 10

        # 将条形图放到产业分析图表 sheet
        chart_sheet.add_chart(bar, "D22")

        # 复制图表数据到产业分析图表 sheet
        for row_idx in range(chart_data_start_row, data_end_row + 1):
            for col_idx in range(1, 4):
                cell = ws.cell(row=row_idx, column=col_idx)
                chart_sheet.cell(row=row_idx, column=col_idx, value=cell.value)
                if cell.number_format:
                    chart_sheet.cell(row=row_idx, column=col_idx).number_format = cell.number_format

        # 设置产业分析图表 sheet 的列宽
        chart_sheet.column_dimensions['A'].width = 30
        chart_sheet.column_dimensions['B'].width = 15
        chart_sheet.column_dimensions['C'].width = 12

    def _get_column_letter(self, col_idx: int) -> str:
        """将列索引转换为 Excel 列字母"""
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result


if __name__ == "__main__":
    # 测试代码
    generator = Sheet2Generator(top_n=10, city_name="重庆市")

    # 模拟数据
    generator.add_data('13', '农副食品加工业', {
        '万州区': 1688, '涪陵区': 2745, '渝中区': 265,
        '沙坪坝区': 619, '九龙坡区': 722, '南岸区': 521
    })

    generator.add_data('14', '食品制造业', {
        '万州区': 1279, '涪陵区': 843, '渝中区': 361,
        '沙坪坝区': 1275, '九龙坡区': 1083, '南岸区': 488
    })

    df = generator.generate()
    print(df)
