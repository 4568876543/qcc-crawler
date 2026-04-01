# Excel操作工具函数
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime

# 样式定义
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')


def _format_filter_text(filters):
    """格式化筛选条件为文本"""
    if not filters:
        return "无"
    parts = []
    if filters.get('地区'):
        parts.append(f"地区={filters['地区']}")
    if filters.get('行业'):
        parts.append(f"行业={filters['行业']}")
    if filters.get('登记状态'):
        parts.append(f"登记状态={filters['登记状态']}")
    if filters.get('制造业'):
        parts.append(f"制造业={filters['制造业']}")
    return " | ".join(parts) if parts else "无"


def save_city_summary_table(output_dir, city_data, city_name="长沙", filters=None):
    """
    保存城市汇总表（制造业各区县企业总数）

    Args:
        output_dir: 输出目录
        city_data: dict, 区县 -> 企业数量
        city_name: 城市名称
        filters: dict, 筛选条件 {'地区': xxx, '行业': xxx, '登记状态': xxx}
    """
    detail_dir = os.path.join(output_dir, "行业明细表")
    os.makedirs(detail_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "区县汇总"

    # 第一行：筛选条件
    filter_text = _format_filter_text(filters)
    ws['A1'] = f"筛选条件: {filter_text}"
    ws['A1'].font = Font(bold=True, size=10, color="666666")
    ws.merge_cells('A1:C1')

    # 第二行：生成时间
    ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=9)

    # 第三行：标题
    ws['A3'] = f"{city_name}市制造业企业数量总表"
    ws['A3'].font = Font(bold=True, size=14)
    ws.merge_cells('A3:C3')

    # 表头（从第4行开始）
    headers = ['区县', '制造业企业总数', '占全市比例']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER
    
    # 按企业数排序
    sorted_data = sorted(city_data.items(), key=lambda x: x[1], reverse=True)
    total_sum = sum(city_data.values())

    row_num = 6
    for district, count in sorted_data:
        ws.cell(row=row_num, column=1, value=district).border = THIN_BORDER
        ws.cell(row=row_num, column=2, value=count).border = THIN_BORDER
        ratio = count / total_sum * 100 if total_sum > 0 else 0
        ws.cell(row=row_num, column=3, value=f"{ratio:.2f}%").border = THIN_BORDER
        row_num += 1

    # 合计行
    ws.cell(row=row_num, column=1, value='合计').font = Font(bold=True)
    ws.cell(row=row_num, column=1).border = THIN_BORDER
    ws.cell(row=row_num, column=2, value=total_sum).font = Font(bold=True)
    ws.cell(row=row_num, column=2).border = THIN_BORDER
    ws.cell(row=row_num, column=3, value='100.00%').border = THIN_BORDER

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12

    # 保存
    output_file = os.path.join(detail_dir, f"00_{city_name}市制造业总表.xlsx")
    wb.save(output_file)
    print(f"[Excel] 城市汇总表已保存: {output_file}")
    return output_file


def save_industry_detail_table(output_dir, industry_code, industry_name, industry_data, city_total_data, city_name="长沙", filters=None):
    """
    保存单个行业的明细表

    Args:
        output_dir: 输出目录
        industry_code: 行业代码
        industry_name: 行业名称
        industry_data: dict, 区县 -> 本行业企业数量
        city_total_data: dict, 区县 -> 制造业企业总数
        city_name: 城市名称
        filters: dict, 筛选条件
    """
    detail_dir = os.path.join(output_dir, "行业明细表")
    os.makedirs(detail_dir, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "区县明细"

    # 第一行：筛选条件
    filter_text = _format_filter_text(filters)
    ws['A1'] = f"筛选条件: {filter_text}"
    ws['A1'].font = Font(bold=True, size=10, color="666666")
    ws.merge_cells('A1:D1')

    # 第二行：生成时间
    ws['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, size=9)

    # 第三行：标题
    ws['A3'] = f"{city_name}市 {industry_name} 企业数量明细表"
    ws['A3'].font = Font(bold=True, size=14)
    ws.merge_cells('A3:D3')

    # 第四行：行业代码
    ws['A4'] = f"行业代码: {industry_code}"
    ws['A4'].font = Font(italic=True, size=9)

    # 表头（从第5行开始）
    headers = ['区县', '本行业企业数', '该区县制造业总数', '占该区县比例']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER

    # 按城市总表的顺序排序
    district_order = sorted(city_total_data.keys(), key=lambda x: city_total_data.get(x, 0), reverse=True)

    total_sum = sum(city_total_data.values())
    industry_sum = sum(industry_data.values())

    row_num = 7  # 数据从第7行开始（第6行是表头）
    for district in district_order:
        count = industry_data.get(district, 0)
        district_total = city_total_data.get(district, 0)
        ratio = count / district_total * 100 if district_total > 0 else 0
        
        ws.cell(row=row_num, column=1, value=district).border = THIN_BORDER
        ws.cell(row=row_num, column=2, value=count).border = THIN_BORDER
        ws.cell(row=row_num, column=3, value=district_total).border = THIN_BORDER
        ws.cell(row=row_num, column=4, value=f"{ratio:.2f}%").border = THIN_BORDER
        row_num += 1
    
    # 合计行
    ws.cell(row=row_num, column=1, value='合计').font = Font(bold=True)
    ws.cell(row=row_num, column=1).fill = TOTAL_FILL
    ws.cell(row=row_num, column=1).border = THIN_BORDER
    ws.cell(row=row_num, column=2, value=industry_sum).font = Font(bold=True)
    ws.cell(row=row_num, column=2).fill = TOTAL_FILL
    ws.cell(row=row_num, column=2).border = THIN_BORDER
    ws.cell(row=row_num, column=3, value=total_sum).font = Font(bold=True)
    ws.cell(row=row_num, column=3).fill = TOTAL_FILL
    ws.cell(row=row_num, column=3).border = THIN_BORDER
    total_ratio = industry_sum / total_sum * 100 if total_sum > 0 else 0
    ws.cell(row=row_num, column=4, value=f"{total_ratio:.2f}%").font = Font(bold=True)
    ws.cell(row=row_num, column=4).fill = TOTAL_FILL
    ws.cell(row=row_num, column=4).border = THIN_BORDER
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    
    # 保存
    safe_name = industry_name.replace('/', '_').replace('\\', '_')
    output_file = os.path.join(detail_dir, f"{industry_code}_{safe_name}.xlsx")
    wb.save(output_file)
    print(f"[Excel] 行业明细表已保存: {industry_code}_{safe_name}.xlsx")
    return output_file


def save_summary_table(output_dir, all_data, city_name="长沙", filters=None):
    """
    保存汇总明细表

    Args:
        output_dir: 输出目录
        all_data: list, 所有数据记录
        city_name: 城市名称
        filters: dict, 筛选条件
    """
    df = pd.DataFrame(all_data)

    # 排除制造业合计行
    df_detail = df[df['行业代码'] != 'C'].copy()

    output_file = os.path.join(output_dir, f"{city_name}制造业企业数量明细表.xlsx")

    wb = Workbook()

    # Sheet1: 明细数据
    ws1 = wb.active
    ws1.title = "明细数据"

    # 第一行：筛选条件
    filter_text = _format_filter_text(filters)
    ws1['A1'] = f"筛选条件: {filter_text}"
    ws1['A1'].font = Font(bold=True, size=10, color="666666")
    ws1.merge_cells('A1:D1')

    # 第二行：生成时间
    ws1['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws1['A2'].font = Font(italic=True, size=9)

    # 第三行：标题
    ws1['A3'] = f"{city_name}市制造业企业数量明细表"
    ws1['A3'].font = Font(bold=True, size=14)
    ws1.merge_cells('A3:D3')

    headers = ['区县', '行业代码', '行业名称', '企业数量']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER

    # 排序
    df_detail['行业代码_num'] = df_detail['行业代码'].astype(int)
    df_detail = df_detail.sort_values(['区县', '行业代码_num'])

    row_num = 6
    for _, row in df_detail.iterrows():
        ws1.cell(row=row_num, column=1, value=row['区县']).border = THIN_BORDER
        ws1.cell(row=row_num, column=2, value=row['行业代码']).border = THIN_BORDER
        ws1.cell(row=row_num, column=3, value=row['行业类别']).border = THIN_BORDER
        ws1.cell(row=row_num, column=4, value=row['企业数量']).border = THIN_BORDER
        row_num += 1

    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 35
    ws1.column_dimensions['D'].width = 12
    
    # Sheet2: 区县汇总
    ws2 = wb.create_sheet(title="区县汇总")
    # 第一行：筛选条件
    ws2['A1'] = f"筛选条件: {filter_text}"
    ws2['A1'].font = Font(bold=True, size=10, color="666666")
    ws2.merge_cells('A1:B1')
    # 第二行：生成时间
    ws2['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws2['A2'].font = Font(italic=True, size=9)
    # 第三行：标题
    ws2['A3'] = f"{city_name}市制造业企业数量 - 区县汇总"
    ws2['A3'].font = Font(bold=True, size=14)

    headers2 = ['区县', '企业总数']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    district_sum = df_detail.groupby('区县')['企业数量'].sum().reset_index()
    district_sum.columns = ['区县', '企业总数']
    district_sum = district_sum.sort_values('企业总数', ascending=False)

    row_num = 6
    for _, row in district_sum.iterrows():
        ws2.cell(row=row_num, column=1, value=row['区县']).border = THIN_BORDER
        ws2.cell(row=row_num, column=2, value=row['企业总数']).border = THIN_BORDER
        row_num += 1

    ws2.cell(row=row_num, column=1, value='合计').font = Font(bold=True)
    ws2.cell(row=row_num, column=1).border = THIN_BORDER
    ws2.cell(row=row_num, column=2, value=df_detail['企业数量'].sum()).font = Font(bold=True)
    ws2.cell(row=row_num, column=2).border = THIN_BORDER

    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 12

    # Sheet3: 行业汇总
    ws3 = wb.create_sheet(title="行业汇总")
    # 第一行：筛选条件
    ws3['A1'] = f"筛选条件: {filter_text}"
    ws3['A1'].font = Font(bold=True, size=10, color="666666")
    ws3.merge_cells('A1:C1')
    # 第二行：生成时间
    ws3['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws3['A2'].font = Font(italic=True, size=9)
    # 第三行：标题
    ws3['A3'] = f"{city_name}市制造业企业数量 - 行业汇总"
    ws3['A3'].font = Font(bold=True, size=14)

    headers3 = ['行业代码', '行业名称', '企业总数']
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=5, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    industry_sum = df_detail.groupby(['行业代码', '行业类别'])['企业数量'].sum().reset_index()
    industry_sum.columns = ['行业代码', '行业名称', '企业总数']
    industry_sum['行业代码_num'] = industry_sum['行业代码'].astype(int)
    industry_sum = industry_sum.sort_values('行业代码_num')

    row_num = 6
    for _, row in industry_sum.iterrows():
        ws3.cell(row=row_num, column=1, value=row['行业代码']).border = THIN_BORDER
        ws3.cell(row=row_num, column=2, value=row['行业名称']).border = THIN_BORDER
        ws3.cell(row=row_num, column=3, value=row['企业总数']).border = THIN_BORDER
        row_num += 1
    
    ws3.cell(row=row_num, column=1, value='C').font = Font(bold=True)
    ws3.cell(row=row_num, column=1).border = THIN_BORDER
    ws3.cell(row=row_num, column=2, value='制造业合计').font = Font(bold=True)
    ws3.cell(row=row_num, column=2).border = THIN_BORDER
    ws3.cell(row=row_num, column=3, value=df_detail['企业数量'].sum()).font = Font(bold=True)
    ws3.cell(row=row_num, column=3).border = THIN_BORDER
    
    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 35
    ws3.column_dimensions['C'].width = 12
    
    wb.save(output_file)
    print(f"[Excel] 汇总明细表已保存: {output_file}")
    return output_file


def validate_data_consistency(city_total_data, all_data, threshold=0.02):
    """
    验证城市汇总表和汇总明细表的数据一致性
    
    Args:
        city_total_data: dict, 城市汇总表数据（区县 -> 企业数量）
        all_data: list, 所有明细数据
        threshold: 允许的误差阈值（默认2%）
    
    Returns:
        dict: 验证结果
    """
    df = pd.DataFrame(all_data)
    df_detail = df[df['行业代码'] != 'C'].copy()
    
    # 计算城市汇总表总额
    city_total = sum(city_total_data.values())
    
    # 计算明细表总额
    detail_total = df_detail['企业数量'].sum()
    
    # 计算误差
    if city_total > 0:
        diff = abs(city_total - detail_total)
        diff_ratio = diff / city_total
    else:
        diff_ratio = 1.0
    
    is_valid = diff_ratio <= threshold
    
    return {
        'is_valid': is_valid,
        'city_total': city_total,
        'detail_total': detail_total,
        'difference': abs(city_total - detail_total),
        'diff_ratio': diff_ratio * 100,  # 转为百分比
        'threshold': threshold * 100
    }


def update_execution_progress(output_file, industry_tasks, completed_count=0, total_count=0):
    """
    更新执行进度表

    Args:
        output_file: Excel文件路径
        industry_tasks: list, 行业任务列表 [{'code': '13', 'name': '农副食品加工业', 'status': 'completed'}, ...]
        completed_count: 已完成数量
        total_count: 总数
    """
    try:
        # 读取现有工作簿
        if os.path.exists(output_file):
            wb = load_workbook(output_file)
        else:
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])

        # 删除旧的执行进度表（如果存在）
        if '执行进度' in wb.sheetnames:
            wb.remove(wb['执行进度'])

        # 创建执行进度表
        ws = wb.create_sheet(title="执行进度", index=0)  # 放在第一个位置

        # 标题
        ws['A1'] = "爬取执行进度监控"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')

        # 生成时间
        ws['A2'] = f"更新时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=9)

        # 统计信息
        total = len(industry_tasks)
        completed = len([t for t in industry_tasks if t.get('status') == 'completed'])
        pending = len([t for t in industry_tasks if t.get('status') == 'pending'])
        in_progress = len([t for t in industry_tasks if t.get('status') == 'in_progress'])
        failed = len([t for t in industry_tasks if t.get('status') == 'failed'])

        ws['A4'] = f"总行业数: {total}"
        ws['A5'] = f"已完成: {completed} | 进行中: {in_progress} | 待处理: {pending} | 失败: {failed}"
        ws['A6'] = f"进度: {completed}/{total} ({completed/total*100:.1f}%)" if total > 0 else "进度: 0/0 (0%)"

        for row in [4, 5, 6]:
            ws[f'A{row}'].font = Font(bold=True)

        # 表头
        headers = ['序号', '行业代码', '行业名称', '状态', '企业数量']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=8, column=col, value=header)
            cell.font = HEADER_FONT_WHITE
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGNMENT
            cell.border = THIN_BORDER

        # 数据行
        status_colors = {
            'completed': 'C6EFCE',  # 绿色
            'in_progress': 'FFEB9C',  # 黄色
            'pending': 'D9D9D9',  # 灰色
            'failed': 'FFC7CE',  # 红色
        }
        status_text_colors = {
            'completed': '006100',
            'in_progress': '9C6500',
            'pending': '666666',
            'failed': '9C0006',
        }

        row_num = 9
        for i, task in enumerate(industry_tasks):
            status = task.get('status', 'pending')
            ws.cell(row=row_num, column=1, value=i+1).border = THIN_BORDER
            ws.cell(row=row_num, column=2, value=task.get('code', '')).border = THIN_BORDER
            ws.cell(row=row_num, column=3, value=task.get('name', '')).border = THIN_BORDER

            status_cell = ws.cell(row=row_num, column=4, value=status)
            status_cell.border = THIN_BORDER
            if status in status_colors:
                status_cell.fill = PatternFill(start_color=status_colors[status], end_color=status_colors[status], fill_type="solid")
                status_cell.font = Font(color=status_text_colors[status])

            ws.cell(row=row_num, column=5, value=task.get('enterprise_count', 0)).border = THIN_BORDER
            row_num += 1

        # 待爬取行业列表（橙色标注）
        pending_tasks = [t for t in industry_tasks if t.get('status') == 'pending']
        if pending_tasks:
            row_num += 1
            ws.cell(row=row_num, column=1, value="▼ 待爬取行业").font = Font(bold=True, color="FF6600")
            ws.merge_cells(f'A{row_num}:E{row_num}')
            row_num += 1
            for task in pending_tasks:
                ws.cell(row=row_num, column=1, value="").border = THIN_BORDER
                ws.cell(row=row_num, column=2, value=task.get('code', '')).border = THIN_BORDER
                ws.cell(row=row_num, column=3, value=task.get('name', '')).border = THIN_BORDER
                ws.cell(row=row_num, column=4, value="待爬取").border = THIN_BORDER
                ws.cell(row=row_num, column=4).fill = PatternFill(start_color='FFE0CC', end_color='FFE0CC', fill_type="solid")
                ws.cell(row=row_num, column=5, value=0).border = THIN_BORDER
                row_num += 1

        # 失败行业列表
        failed_tasks = [t for t in industry_tasks if t.get('status') == 'failed']
        if failed_tasks:
            row_num += 1
            ws.cell(row=row_num, column=1, value="✗ 失败行业").font = Font(bold=True, color="CC0000")
            ws.merge_cells(f'A{row_num}:E{row_num}')
            row_num += 1
            for task in failed_tasks:
                ws.cell(row=row_num, column=1, value="").border = THIN_BORDER
                ws.cell(row=row_num, column=2, value=task.get('code', '')).border = THIN_BORDER
                ws.cell(row=row_num, column=3, value=task.get('name', '')).border = THIN_BORDER
                ws.cell(row=row_num, column=4, value="失败").border = THIN_BORDER
                ws.cell(row=row_num, column=4).fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type="solid")
                ws.cell(row=row_num, column=4).font = Font(color='9C0006')
                ws.cell(row=row_num, column=5, value=task.get('error_message', '')).border = THIN_BORDER
                row_num += 1

        # 设置列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15

        # 保存
        wb.save(output_file)
        print(f"[Excel] 执行进度表已更新: {output_file}")
        return True

    except Exception as e:
        print(f"[Excel] 更新执行进度表失败: {e}")
        return False


def create_excel_template(output_file):
    """
    创建Excel模板文件，包含总览表结构
    
    Args:
        output_file (str): 输出的Excel文件路径
    """
    # 确保输出目录存在
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    
    # 创建工作簿
    wb = Workbook()
    
    # 删除默认的Sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 创建总览表
    overview_sheet = wb.create_sheet(title="总览")
    overview_sheet['A1'] = "重庆市制造业企业统计"
    overview_sheet['A1'].font = Font(bold=True, size=14)
    overview_sheet.merge_cells('A1:C1')
    
    # 设置表头
    headers = ["区县", "行业类别", "企业数量"]
    for col, header in enumerate(headers, 1):
        cell = overview_sheet.cell(row=3, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER
    
    # 调整列宽
    overview_sheet.column_dimensions['A'].width = 20
    overview_sheet.column_dimensions['B'].width = 35
    overview_sheet.column_dimensions['C'].width = 12
    
    # 保存工作簿
    wb.save(output_file)
    print(f"[Excel] 模板已创建: {output_file}")


def update_excel_data(output_file, data):
    """
    更新Excel文件中的总览表数据
    
    Args:
        output_file (str): Excel文件路径
        data (list): 包含区县、行业类别和企业数量的数据列表
    """
    # 将数据转换为DataFrame
    new_data = pd.DataFrame(data, columns=["区县", "行业类别", "企业数量"])
    
    # 检查文件是否存在
    if os.path.exists(output_file):
        try:
            # 读取现有数据
            existing_data = pd.read_excel(output_file, sheet_name="总览", header=2)
            
            # 合并数据，保留最新的企业数量
            combined_data = pd.concat([existing_data, new_data], ignore_index=True)
            combined_data = combined_data.drop_duplicates(subset=["区县", "行业类别"], keep="last")
        except Exception as e:
            print(f"[Excel] 读取现有数据失败: {e}")
            combined_data = new_data
    else:
        combined_data = new_data
    
    # 按区县和行业类别排序
    combined_data = combined_data.sort_values(by=["区县", "行业类别"]).reset_index(drop=True)
    
    # 使用ExcelWriter写入
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        combined_data.to_excel(writer, sheet_name="总览", index=False, startrow=2, header=True)
    
    print(f"[Excel] 总览表已更新: {len(combined_data)} 条记录")


def create_district_sheets(output_file, districts, industries):
    """
    为每个区县创建单独的工作表
    
    Args:
        output_file (str): Excel文件路径
        districts (list): 区县列表
        industries (dict): 行业代码和名称字典
    """
    if not os.path.exists(output_file):
        create_excel_template(output_file)
    
    # 读取现有Excel
    with pd.ExcelFile(output_file) as xl:
        existing_sheets = xl.sheet_names
    
    # 使用ExcelWriter追加工作表
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        # 复制现有工作表（跳过将创建的区县工作表）
        for sheet_name in existing_sheets:
            if sheet_name not in districts and sheet_name not in ["区县汇总", "行业汇总"]:
                df = pd.read_excel(output_file, sheet_name=sheet_name)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 为每个区县创建工作表
        for district in districts:
            # 创建区县数据表
            district_df = pd.DataFrame({
                "行业代码": list(industries.keys()),
                "行业名称": list(industries.values()),
                "企业数量": [0] * len(industries)
            })
            district_df.to_excel(writer, sheet_name=district, index=False)
    
    print(f"[Excel] 已创建 {len(districts)} 个区县工作表")


def update_district_sheet(output_file, district, industry_code, industry_name, count):
    """
    更新区县工作表中的数据
    
    Args:
        output_file (str): Excel文件路径
        district (str): 区县名称
        industry_code (str): 行业代码
        industry_name (str): 行业名称
        count (int): 企业数量
    """
    if not os.path.exists(output_file):
        print(f"[Excel] 文件不存在: {output_file}")
        return
    
    try:
        # 读取所有工作表
        all_sheets = {}
        with pd.ExcelFile(output_file) as xl:
            for sheet_name in xl.sheet_names:
                all_sheets[sheet_name] = pd.read_excel(xl, sheet_name=sheet_name)
        
        # 更新区县工作表
        if district in all_sheets:
            district_df = all_sheets[district]
            
            # 确保行业代码为字符串类型进行比较
            district_df["行业代码"] = district_df["行业代码"].astype(str)
            
            # 更新数据
            mask = district_df["行业代码"] == str(industry_code)
            if mask.any():
                district_df.loc[mask, "企业数量"] = count
            else:
                # 添加新行
                new_row = pd.DataFrame({
                    "行业代码": [str(industry_code)],
                    "行业名称": [industry_name],
                    "企业数量": [count]
                })
                district_df = pd.concat([district_df, new_row], ignore_index=True)
            
            all_sheets[district] = district_df
        
        # 写回所有工作表
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
            for sheet_name, df in all_sheets.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        print(f"[Excel] 更新 {district} - {industry_name}: {count} 家企业")
        
    except Exception as e:
        print(f"[Excel] 更新区县工作表失败: {e}")


def create_summary_sheet(output_file, districts, industries):
    """
    创建汇总工作表
    
    Args:
        output_file (str): Excel文件路径
        districts (list): 区县列表
        industries (dict): 行业代码和名称字典
    """
    if not os.path.exists(output_file):
        create_excel_template(output_file)
    
    # 创建按区县汇总的DataFrame
    district_summary = pd.DataFrame({
        "区县": districts,
        "企业总数": [0] * len(districts)
    })
    
    # 创建按行业汇总的DataFrame
    industry_summary = pd.DataFrame({
        "行业代码": list(industries.keys()),
        "行业名称": list(industries.values()),
        "企业总数": [0] * len(industries)
    })
    
    # 读取现有工作表
    all_sheets = {}
    if os.path.exists(output_file):
        with pd.ExcelFile(output_file) as xl:
            for sheet_name in xl.sheet_names:
                if sheet_name not in ["区县汇总", "行业汇总"]:
                    all_sheets[sheet_name] = pd.read_excel(xl, sheet_name=sheet_name)
    
    # 写入所有工作表
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
        # 写入原有工作表
        for sheet_name, df in all_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 写入汇总工作表
        district_summary.to_excel(writer, sheet_name="区县汇总", index=False)
        industry_summary.to_excel(writer, sheet_name="行业汇总", index=False)
    
    print(f"[Excel] 汇总工作表已创建")


def update_summary_sheet(output_file, data):
    """
    更新汇总工作表
    
    Args:
        output_file (str): Excel文件路径
        data (list): 包含区县、行业类别和企业数量的数据列表
    """
    if not os.path.exists(output_file):
        print(f"[Excel] 文件不存在: {output_file}")
        return
    
    if not data:
        print(f"[Excel] 无数据可更新汇总表")
        return
    
    # 将数据转换为DataFrame
    df = pd.DataFrame(data)
    
    # 按区县汇总
    district_summary = df.groupby("区县")["企业数量"].sum().reset_index()
    district_summary.columns = ["区县", "企业总数"]
    
    # 按行业汇总
    industry_summary = df.groupby(["行业代码", "行业类别"])["企业数量"].sum().reset_index()
    industry_summary.columns = ["行业代码", "行业名称", "企业总数"]
    
    # 读取现有工作表
    all_sheets = {}
    with pd.ExcelFile(output_file) as xl:
        for sheet_name in xl.sheet_names:
            if sheet_name not in ["区县汇总", "行业汇总"]:
                all_sheets[sheet_name] = pd.read_excel(xl, sheet_name=sheet_name)
    
    # 写入所有工作表
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
        # 写入原有工作表
        for sheet_name, df_sheet in all_sheets.items():
            df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 写入更新后的汇总工作表
        district_summary.to_excel(writer, sheet_name="区县汇总", index=False)
        industry_summary.to_excel(writer, sheet_name="行业汇总", index=False)
    
    print(f"[Excel] 汇总表已更新: 区县 {len(district_summary)} 条, 行业 {len(industry_summary)} 条")


def update_all_district_sheets(output_file, data, districts, industries):
    """
    批量更新所有区县工作表
    
    Args:
        output_file (str): Excel文件路径
        data (list): 完整数据列表
        districts (list): 区县列表
        industries (dict): 行业字典
    """
    if not data:
        return
    
    df = pd.DataFrame(data)
    
    # 确保行业代码为字符串类型
    df["行业代码"] = df["行业代码"].astype(str)
    
    # 读取现有工作表
    all_sheets = {}
    with pd.ExcelFile(output_file) as xl:
        for sheet_name in xl.sheet_names:
            all_sheets[sheet_name] = pd.read_excel(xl, sheet_name=sheet_name)
    
    # 更新每个区县工作表
    for district in districts:
        if district in all_sheets:
            district_df = all_sheets[district]
            
            # 确保行业代码为字符串类型进行比较
            district_df["行业代码"] = district_df["行业代码"].astype(str)
            
            # 获取该区县的数据
            district_data = df[df["区县"] == district]
            
            for _, row in district_data.iterrows():
                mask = district_df["行业代码"] == str(row["行业代码"])
                if mask.any():
                    district_df.loc[mask, "企业数量"] = row["企业数量"]
            
            all_sheets[district] = district_df
    
    # 写回所有工作表
    with pd.ExcelWriter(output_file, engine='openpyxl', mode='w') as writer:
        for sheet_name, df_sheet in all_sheets.items():
            df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)
    
    print(f"[Excel] 已批量更新所有区县工作表")
